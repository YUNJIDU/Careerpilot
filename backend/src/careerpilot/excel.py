import os
from collections.abc import Iterable, Mapping
from datetime import date, datetime
from enum import StrEnum
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
from uuid import UUID, uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ConfigDict, Field, ValidationError

from careerpilot.contracts import Checkpoint, Job, JobStatus
from careerpilot.security import escape_excel_formula, safe_path

SCHEMA_VERSION = "1.0"
SHEET_NAME = "Tracker"
COLUMNS = (
    "投递时间",
    "公司名称",
    "岗位",
    "简历通过",
    "测评",
    "笔试",
    "一面",
    "二面",
    "三面",
    "HR 面",
    "终面",
    "当前阶段",
    "截止时间",
    "JD 链接",
    "最近更新时间",
    "备注",
)
RESUME_COLUMN = "当前简历"
_ID_COLUMN = "_application_id"
_VERSION_COLUMN = "_row_version"
_MAX_CELL_LENGTH = 5000
_MAX_FILE_SIZE = 20 * 1024 * 1024


class ExcelError(ValueError):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        sheet: str | None = None,
        row: int | None = None,
        column: str | None = None,
    ) -> None:
        self.code, self.sheet, self.row, self.column = code, sheet, row, column
        location = "/".join(str(value) for value in (sheet, row, column) if value is not None)
        super().__init__(f"{message}{f' [{location}]' if location else ''}")


class TrackerRow(BaseModel):
    model_config = ConfigDict(extra="forbid")
    application_id: UUID = Field(default_factory=uuid4)
    row_version: int = Field(default=1, ge=1)
    values: dict[str, str | date | datetime | None]
    generated_id: bool = False
    resume_column_present: bool = False


class ChangeKind(StrEnum):
    CREATE = "create"
    UPDATE = "update"
    CLEAR = "clear"
    NOOP = "noop"
    CONFLICT = "conflict"


class DiffCommand(BaseModel):
    application_id: UUID
    field: str | None = None
    kind: ChangeKind
    old_value: Any = None
    new_value: Any = None
    system_value: Any = None
    row_version: int


def _check_path(path: Path, root: Path | None) -> Path:
    if path.suffix.lower() != ".xlsx":
        raise ExcelError("excel.unsupported_type", "only .xlsx files are supported")
    resolved = safe_path(root, path) if root else path.resolve()
    if resolved.exists() and resolved.stat().st_size > _MAX_FILE_SIZE:
        raise ExcelError("excel.too_large", "workbook exceeds 20 MiB")
    return resolved


def write_tracker(path: Path, rows: Iterable[TrackerRow], root: Path | None = None) -> Path:
    target = _check_path(path, root)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    headers = (*COLUMNS, _ID_COLUMN, _VERSION_COLUMN, RESUME_COLUMN)
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for index, name in enumerate(COLUMNS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = max(12, min(30, len(name) + 6))
    sheet.column_dimensions[get_column_letter(len(COLUMNS) + 1)].hidden = True
    sheet.column_dimensions[get_column_letter(len(COLUMNS) + 2)].hidden = True
    sheet.column_dimensions[get_column_letter(len(COLUMNS) + 3)].width = 24

    for row in rows:
        cells: list[Any] = []
        for column in COLUMNS:
            value = row.values.get(column)
            if isinstance(value, str):
                value = escape_excel_formula(value)
            cells.append(value)
        sheet.append(
            (*cells, str(row.application_id), row.row_version, row.values.get(RESUME_COLUMN))
        )
    for column in ("A", "M", "O"):
        for cell in sheet[column][1:]:
            cell.number_format = "yyyy-mm-dd"

    metadata = workbook.create_sheet("_meta")
    metadata.sheet_state = "hidden"
    metadata.append(("schema_version", SCHEMA_VERSION))

    with NamedTemporaryFile(suffix=".xlsx", dir=target.parent, delete=False) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        read_tracker(temporary)
        os.replace(temporary, target)
    finally:
        temporary.unlink(missing_ok=True)
    return target


def read_tracker(path: Path, root: Path | None = None) -> list[TrackerRow]:
    source = _check_path(path, root)
    try:
        workbook = load_workbook(source, data_only=False, read_only=False)
    except Exception as exc:
        raise ExcelError("excel.corrupt", "workbook cannot be opened") from exc
    if SHEET_NAME not in workbook.sheetnames:
        raise ExcelError("excel.missing_sheet", "Tracker sheet is missing", sheet=SHEET_NAME)
    if "_meta" not in workbook.sheetnames:
        raise ExcelError("excel.missing_metadata", "metadata sheet is missing")
    version = workbook["_meta"]["B1"].value
    if version != SCHEMA_VERSION:
        raise ExcelError("excel.unknown_version", f"unsupported schema version: {version}")
    if getattr(workbook, "_external_links", []):
        raise ExcelError("excel.external_links", "external links are not allowed")

    sheet = workbook[SHEET_NAME]
    headers = [cell.value for cell in sheet[1]]
    expected = [*COLUMNS, _ID_COLUMN, _VERSION_COLUMN]
    duplicates = {name for name in headers if name is not None and headers.count(name) > 1}
    if duplicates:
        raise ExcelError(
            "excel.duplicate_column", f"duplicate columns: {sorted(duplicates)}", sheet=SHEET_NAME
        )
    missing = [name for name in expected if name not in headers]
    if missing:
        raise ExcelError("excel.missing_column", f"missing columns: {missing}", sheet=SHEET_NAME)

    positions = {name: headers.index(name) + 1 for name in expected}
    result: list[TrackerRow] = []
    seen: set[UUID] = set()
    for row_index in range(2, sheet.max_row + 1):
        resume_column_present = RESUME_COLUMN in headers
        visible_columns = (*COLUMNS, RESUME_COLUMN) if resume_column_present else COLUMNS
        if all(
            sheet.cell(row_index, headers.index(name) + 1).value is None for name in visible_columns
        ):
            continue
        raw_id = sheet.cell(row_index, positions[_ID_COLUMN]).value
        generated_id = raw_id in (None, "")
        if generated_id:
            application_id = uuid4()
        else:
            try:
                application_id = UUID(str(raw_id))
            except (TypeError, ValueError) as exc:
                raise ExcelError(
                    "excel.invalid_id",
                    "application ID is invalid",
                    sheet=SHEET_NAME,
                    row=row_index,
                    column=_ID_COLUMN,
                ) from exc
        if application_id in seen:
            raise ExcelError(
                "excel.duplicate_id",
                "application ID is duplicated",
                sheet=SHEET_NAME,
                row=row_index,
                column=_ID_COLUMN,
            )
        seen.add(application_id)
        values: dict[str, Any] = {}
        for name in COLUMNS:
            value = sheet.cell(row_index, positions[name]).value
            if isinstance(value, str) and len(value) > _MAX_CELL_LENGTH:
                raise ExcelError(
                    "excel.value_too_long",
                    "cell exceeds maximum length",
                    sheet=SHEET_NAME,
                    row=row_index,
                    column=name,
                )
            if isinstance(value, str) and value.startswith("="):
                raise ExcelError(
                    "excel.formula",
                    "formula cells are not accepted",
                    sheet=SHEET_NAME,
                    row=row_index,
                    column=name,
                )
            if (
                name in {"投递时间", "截止时间", "最近更新时间"}
                and value is not None
                and not isinstance(value, (date, datetime))
            ):
                raise ExcelError(
                    "excel.invalid_date",
                    "date cell must contain an Excel date",
                    sheet=SHEET_NAME,
                    row=row_index,
                    column=name,
                )
            if name in {"投递时间", "截止时间"} and isinstance(value, datetime):
                value = value.date()
            values[name] = value
        if resume_column_present:
            resume_value = sheet.cell(row_index, headers.index(RESUME_COLUMN) + 1).value
            if resume_value is not None and not isinstance(resume_value, str):
                raise ExcelError(
                    "excel.invalid_resume",
                    "current resume must be text",
                    sheet=SHEET_NAME,
                    row=row_index,
                    column=RESUME_COLUMN,
                )
            values[RESUME_COLUMN] = resume_value.strip() if resume_value else None
        try:
            result.append(
                TrackerRow(
                    application_id=application_id,
                    row_version=sheet.cell(row_index, positions[_VERSION_COLUMN]).value or 1,
                    values=values,
                    generated_id=generated_id,
                    resume_column_present=resume_column_present,
                )
            )
        except ValidationError as exc:
            raise ExcelError(
                "excel.invalid_row_version",
                "row version must be a positive integer",
                sheet=SHEET_NAME,
                row=row_index,
                column=_VERSION_COLUMN,
            ) from exc
    return result


def diff_tracker(
    excel_rows: Iterable[TrackerRow],
    applications: Mapping[UUID, Mapping[str, Any]],
    baseline: Mapping[UUID, Mapping[str, Any]],
) -> list[DiffCommand]:
    commands: list[DiffCommand] = []
    for row in excel_rows:
        current = applications.get(row.application_id)
        if current is None:
            commands.append(
                DiffCommand(
                    application_id=row.application_id,
                    kind=ChangeKind.CREATE,
                    new_value=row.values,
                    row_version=row.row_version,
                )
            )
            continue
        previous = baseline.get(row.application_id, {})
        for field in COLUMNS:
            excel_value = row.values.get(field)
            system_value = current.get(field)
            old_value = previous.get(field)
            if excel_value == system_value:
                kind = ChangeKind.NOOP
            elif system_value == old_value:
                kind = ChangeKind.CLEAR if excel_value in (None, "") else ChangeKind.UPDATE
            elif excel_value == old_value:
                kind = ChangeKind.UPDATE
            else:
                kind = ChangeKind.CONFLICT
            commands.append(
                DiffCommand(
                    application_id=row.application_id,
                    field=field,
                    kind=kind,
                    old_value=old_value,
                    new_value=excel_value,
                    system_value=system_value,
                    row_version=row.row_version,
                )
            )
    return commands


class MemoryJobStore:
    def __init__(self) -> None:
        self.jobs: dict[UUID, Job] = {}

    def start(self, idempotency_key: str) -> Job:
        existing = next(
            (job for job in self.jobs.values() if job.idempotency_key == idempotency_key),
            None,
        )
        if existing:
            return existing
        job = Job(job_type="excel_sync", idempotency_key=idempotency_key, status=JobStatus.RUNNING)
        self.jobs[job.job_id] = job
        return job

    def checkpoint(self, job_id: UUID, step: str, payload: dict[str, Any] | None = None) -> Job:
        job = self.jobs[job_id]
        job.current_step = step
        job.completed_steps.append(step)
        job.checkpoint = Checkpoint(step=step, payload=payload or {})
        return job
