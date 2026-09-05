from datetime import UTC, datetime
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from careerpilot.api import create_app
from careerpilot.core import Database
from careerpilot.excel import COLUMNS, write_tracker
from careerpilot.mail import MailItem, MailSyncService


class Inbox:
    def __init__(self, *items):
        self.items = list(items)

    def fetch(self):
        return self.items


def message(key, day, role="Engineer", extra="阶段：一面"):
    return MailItem(
        f"<{key}@example>",
        "jobs@example.com",
        "面试通知",
        datetime(2026, 7, day, tzinfo=UTC),
        "公司：Acme\n" + (f"岗位：{role}\n" if role else "") + extra,
        key * 64,
    )


def test_deleted_application_is_not_resurrected_by_old_mail(tmp_path: Path):
    db = Database(tmp_path / "careerpilot.db")
    service = MailSyncService(db)
    tracker = tmp_path / "tracker.xlsx"
    inbox = Inbox(message("a", 1))
    service.sync(inbox, "fixture", tracker, "first")
    write_tracker(tracker, [])
    service.sync(inbox, "fixture", tracker, "second")
    assert service.applications.list() == []


def test_new_mail_changes_only_new_fields_and_old_mail_cannot_rewind(tmp_path: Path):
    db = Database(tmp_path / "careerpilot.db")
    service = MailSyncService(db)
    tracker = tmp_path / "tracker.xlsx"
    service.sync(
        Inbox(message("a", 2, extra="阶段：二面\n截止时间：2026-08-12")),
        "fixture",
        tracker,
        "first",
    )
    app = service.applications.list()[0]
    service.applications.apply_field_change(
        app.application_id, "备注", "短信人工补充", source="user", idempotency_key="note"
    )
    service.excel.export_workbook(tracker)
    service.sync(
        Inbox(message("b", 1, extra="阶段：一面\n截止时间：2026-08-01")), "fixture", tracker, "old"
    )
    current = service.applications.get(app.application_id)
    assert current.values["当前阶段"] == "二面"
    assert str(current.values["截止时间"]) == "2026-08-12"
    assert current.values["备注"] == "短信人工补充"
    service.sync(Inbox(message("c", 3, extra="截止时间：2026-08-15")), "fixture", tracker, "new")
    updated = service.applications.get(app.application_id)
    assert str(updated.values["截止时间"]) == "2026-08-15"
    assert updated.values["当前阶段"] == "二面"
    assert updated.values["备注"] == "短信人工补充"
    version = updated.version
    service.sync(Inbox(message("c", 3, extra="截止时间：2026-08-15")), "fixture", tracker, "repeat")
    assert service.applications.get(app.application_id).version == version


def test_ambiguous_company_mail_is_not_assigned_to_latest_role(tmp_path: Path):
    service = MailSyncService(Database(tmp_path / "careerpilot.db"))
    service.sync(
        Inbox(
            message("a", 1, "One", "阶段：已投递"),
            message("b", 2, "Two", "阶段：已投递"),
            message("c", 3, ""),
        ),
        "fixture",
        tmp_path / "tracker.xlsx",
        "batch",
    )
    assert all(a.values["当前阶段"] == "已投递" for a in service.applications.list())
    assert len(service.emails.linked()) == 2


def test_web_edit_does_not_overwrite_unimported_excel(tmp_path: Path):
    client = TestClient(create_app(data_dir=tmp_path))
    app = client.post(
        "/api/v1/applications",
        json={"company": "Acme", "role": "Engineer", "idempotency_key": "create"},
    ).json()
    tracker = tmp_path / "tracker.xlsx"
    book = load_workbook(tracker)
    book.active.cell(2, COLUMNS.index("备注") + 1, "SMS edit")
    book.save(tracker)
    before = tracker.read_bytes()
    response = client.patch(
        f"/api/v1/applications/{app['application_id']}",
        json={
            "changes": {"岗位": "Changed"},
            "expected_version": app["version"],
            "idempotency_key": "patch",
        },
    )
    assert response.status_code == 409
    assert tracker.read_bytes() == before
    assert client.get(f"/api/v1/applications/{app['application_id']}").json()["role"] == "Engineer"


def test_manual_stage_conflict_is_visible_and_preserved(tmp_path: Path):
    service = MailSyncService(Database(tmp_path / "careerpilot.db"))
    tracker = tmp_path / "tracker.xlsx"
    service.sync(Inbox(message("a", 1)), "fixture", tracker, "first")
    app = service.applications.list()[0]
    service.applications.apply_field_change(
        app.application_id, "当前阶段", "短信二面", source="user", idempotency_key="sms"
    )
    service.excel.export_workbook(tracker)
    result = service.sync(Inbox(message("b", 2, extra="阶段：已拒绝")), "fixture", tracker, "next")
    assert service.applications.get(app.application_id).values["当前阶段"] == "短信二面"
    assert result["conflicts"] >= 1
    assert any(
        p["source"] == "mail_conflict"
        for p in service.applications.details(app.application_id)["provenance"]
    )


def test_interrupted_sync_recovers_without_importing_stale_excel(tmp_path: Path):
    import pytest

    from careerpilot.mail import MailSyncError

    class Interrupted:
        def fetch(self):
            yield message("b", 2, "Second")
            raise OSError("connection lost")

    tracker = tmp_path / "tracker.xlsx"
    service = MailSyncService(Database(tmp_path / "careerpilot.db"))
    service.sync(Inbox(message("a", 1)), "fixture", tracker, "first")
    with pytest.raises(MailSyncError):
        service.sync(Interrupted(), "fixture", tracker, "interrupted")
    service = MailSyncService(Database(tmp_path / "careerpilot.db"))
    service.sync(Inbox(message("b", 2, "Second")), "fixture", tracker, "retry")
    assert {a.role for a in service.applications.list()} == {"Engineer", "Second"}
    service.excel.import_workbook(tracker, "verify")
    assert len(service.applications.list()) == 2


def test_failed_export_preserves_updates_until_retry(tmp_path: Path, monkeypatch):
    import pytest

    from careerpilot.mail import MailSyncError

    tracker = tmp_path / "tracker.xlsx"
    service = MailSyncService(Database(tmp_path / "careerpilot.db"))
    service.sync(Inbox(message("a", 1)), "fixture", tracker, "first")
    export = service.excel.export_workbook

    def locked(*args):
        raise PermissionError("workbook locked")

    monkeypatch.setattr(service.excel, "export_workbook", locked)
    with pytest.raises(MailSyncError):
        service.sync(Inbox(message("b", 2, extra="阶段：二面")), "fixture", tracker, "locked")
    monkeypatch.setattr(service.excel, "export_workbook", export)
    service.sync(Inbox(), "fixture", tracker, "retry")
    service.excel.import_workbook(tracker, "verify")
    assert service.applications.list()[0].values["当前阶段"] == "二面"


def test_pending_sync_blocks_excel_import_and_external_changes(tmp_path: Path, monkeypatch):
    import pytest

    from careerpilot.mail import MailSyncError

    tracker = tmp_path / "tracker.xlsx"
    service = MailSyncService(Database(tmp_path / "careerpilot.db"))
    service.sync(Inbox(message("a", 1)), "fixture", tracker, "first")

    def locked(*args):
        raise PermissionError("locked")

    monkeypatch.setattr(service.excel, "export_workbook", locked)
    with pytest.raises(MailSyncError):
        service.sync(Inbox(message("b", 2, "Second")), "fixture", tracker, "fail")
    with pytest.raises(ValueError):
        service.excel.import_workbook(tracker, "unsafe-import")
    with pytest.raises(ValueError):
        service.excel.require_unmodified(tracker)
    write_tracker(tracker, [])
    with pytest.raises(MailSyncError):
        service.sync(Inbox(), "fixture", tracker, "changed")
    assert len(service.applications.list()) == 2


def test_unresolved_mail_can_be_associated_without_overwriting_fields(tmp_path: Path):
    import pytest

    service = MailSyncService(Database(tmp_path / "careerpilot.db"))
    tracker = tmp_path / "tracker.xlsx"
    service.sync(
        Inbox(message("a", 1, "One"), message("b", 1, "Two"), message("c", 2, "", "阶段：二面")),
        "fixture",
        tracker,
        "sync",
    )
    unresolved = service.emails.unresolved()
    assert len(unresolved) == 1
    app = service.applications.list()[0]
    service.emails.associate(unresolved[0]["email_id"], app.application_id)
    assert service.emails.unresolved() == []
    assert service.applications.get(app.application_id).values == app.values
    write_tracker(tracker, [])
    service.excel.import_workbook(tracker, "delete")
    replacement = service.applications.create("Acme", "New", idempotency_key="replacement")
    with pytest.raises(ValueError):
        service.emails.associate(unresolved[0]["email_id"], replacement.application_id)


def test_web_export_failure_can_recover_without_losing_saved_edit(tmp_path: Path, monkeypatch):
    from careerpilot.core import ExcelSyncService

    client = TestClient(create_app(data_dir=tmp_path))
    app = client.post(
        "/api/v1/applications",
        json={"company": "Acme", "role": "Engineer", "idempotency_key": "create"},
    ).json()
    export = ExcelSyncService.export_workbook

    def locked(*args):
        raise PermissionError("locked")

    monkeypatch.setattr(ExcelSyncService, "export_workbook", locked)
    response = client.patch(
        f"/api/v1/applications/{app['application_id']}",
        json={
            "changes": {"备注": "saved note"},
            "expected_version": app["version"],
            "idempotency_key": "patch",
        },
    )
    assert response.status_code == 500
    monkeypatch.setattr(ExcelSyncService, "export_workbook", export)
    assert (
        client.post(
            "/api/v1/excel-sync-jobs",
            json={
                "path": "tracker.xlsx",
                "direction": "import",
                "idempotency_key": "unsafe",
                "destructive_confirmed": True,
            },
        ).status_code
        == 422
    )
    assert (
        client.post(
            "/api/v1/excel-sync-jobs",
            json={"path": "tracker.xlsx", "direction": "export", "idempotency_key": "recover"},
        ).status_code
        == 200
    )
    assert (
        client.post(
            "/api/v1/excel-sync-jobs",
            json={
                "path": "tracker.xlsx",
                "direction": "import",
                "idempotency_key": "verify",
                "destructive_confirmed": True,
            },
        ).status_code
        == 200
    )
    assert (
        client.get(f"/api/v1/applications/{app['application_id']}").json()["values"]["备注"]
        == "saved note"
    )
