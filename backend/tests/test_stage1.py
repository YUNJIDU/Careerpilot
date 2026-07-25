from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from careerpilot.excel import (
    COLUMNS,
    ChangeKind,
    ExcelError,
    MemoryJobStore,
    TrackerRow,
    diff_tracker,
    read_tracker,
    write_tracker,
)


def sample_row(**values: object) -> TrackerRow:
    defaults = dict.fromkeys(COLUMNS)
    defaults.update({"公司名称": "Acme", "岗位": "Engineer", **values})
    return TrackerRow(values=defaults)


def test_tracker_round_trip_and_formula_escape(tmp_path: Path) -> None:
    path = tmp_path / "tracker.xlsx"
    original = sample_row(备注="=HYPERLINK(\"bad\")")
    write_tracker(path, [original])
    [loaded] = read_tracker(path)
    assert loaded.application_id == original.application_id
    assert loaded.values["备注"].startswith("'=")
    workbook = load_workbook(path)
    assert workbook["Tracker"].freeze_panes == "A2"
    assert workbook["_meta"].sheet_state == "hidden"


def test_reader_locates_missing_column_and_duplicate_id(tmp_path: Path) -> None:
    path = write_tracker(tmp_path / "tracker.xlsx", [sample_row(), sample_row()])
    workbook = load_workbook(path)
    workbook["Tracker"]["A1"] = "错误列"
    workbook.save(path)
    with pytest.raises(ExcelError, match="missing columns") as missing:
        read_tracker(path)
    assert missing.value.sheet == "Tracker"

    first, second = sample_row(), sample_row()
    path = write_tracker(tmp_path / "duplicates.xlsx", [first, second])
    workbook = load_workbook(path)
    workbook["Tracker"]["Q3"] = str(first.application_id)
    workbook.save(path)
    with pytest.raises(ExcelError) as duplicate:
        read_tracker(path)
    assert duplicate.value.code == "excel.duplicate_id"
    assert duplicate.value.row == 3


def test_reader_rejects_formula_corruption_and_escape(tmp_path: Path) -> None:
    path = write_tracker(tmp_path / "tracker.xlsx", [sample_row()])
    workbook = load_workbook(path)
    workbook["Tracker"]["P2"] = "=1+1"
    workbook.save(path)
    with pytest.raises(ExcelError, match="formula"):
        read_tracker(path)
    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not a workbook")
    with pytest.raises(ExcelError) as error:
        read_tracker(corrupt)
    assert error.value.code == "excel.corrupt"
    with pytest.raises(ValueError, match="escapes"):
        read_tracker(tmp_path / ".." / "outside.xlsx", root=tmp_path)


def test_reader_locates_invalid_date(tmp_path: Path) -> None:
    path = write_tracker(tmp_path / "tracker.xlsx", [sample_row()])
    workbook = load_workbook(path)
    workbook["Tracker"]["A2"] = "tomorrow-ish"
    workbook.save(path)
    with pytest.raises(ExcelError) as error:
        read_tracker(path)
    assert (error.value.code, error.value.row, error.value.column) == (
        "excel.invalid_date",
        2,
        "投递时间",
    )


def test_diff_covers_all_commands() -> None:
    app_id = uuid4()
    create = sample_row()
    row = sample_row()
    row.application_id = app_id
    row.values.update(
        {
            "公司名称": "Excel edit",
            "岗位": None,
            "当前阶段": "same",
            "备注": "Excel conflict",
        }
    )
    current = dict.fromkeys(COLUMNS)
    current.update(
        {"公司名称": "old", "岗位": "old", "当前阶段": "same", "备注": "System conflict"}
    )
    baseline = dict(current)
    baseline["备注"] = "old"
    commands = diff_tracker([create, row], {app_id: current}, {app_id: baseline})
    kinds = {command.kind for command in commands}
    assert {
        ChangeKind.CREATE,
        ChangeKind.UPDATE,
        ChangeKind.CLEAR,
        ChangeKind.NOOP,
        ChangeKind.CONFLICT,
    } <= kinds


def test_atomic_write_preserves_existing_file_on_validation_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    path = tmp_path / "tracker.xlsx"
    path.write_bytes(b"original")
    monkeypatch.setattr("careerpilot.excel.read_tracker", lambda _: (_ for _ in ()).throw(ValueError()))
    with pytest.raises(ValueError):
        write_tracker(path, [sample_row()])
    assert path.read_bytes() == b"original"


def test_memory_job_is_idempotent_and_checkpoints() -> None:
    store = MemoryJobStore()
    first = store.start("same-input")
    assert store.start("same-input").job_id == first.job_id
    assert store.checkpoint(first.job_id, "validated").completed_steps == ["validated"]
